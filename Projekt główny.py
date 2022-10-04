from multiprocessing import Event
from unicodedata import name
from PyQt5.QtWidgets import * 
from PyQt5.uic import loadUi
import sys
from ustawienia import Ui_Ustawienia
from PyQt5 import QtCore, QtGui, QtWidgets
from okienko import Ui_Form as Ui_okienko
from wyniki import Ui_Wyniki
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from wzorwynikow import Ui_Wzor_wynikow
import openpyxl
from cmath import exp
import math
import cmath
import numpy as np
import matplotlib.pyplot as plt


class MainUI(QMainWindow):
    def __init__(self):
        super(MainUI, self).__init__()
        loadUi('menu.ui', self)
        self.Wybor.triggered.connect(lambda: self.openWindow())
        self.Dodawanie_temperatury.clicked.connect(lambda: self.dodaj_temperautre())
        self.tabWidget.tabCloseRequested.connect(lambda index: self.closeTab(index))
        self.odswiez_przycisk.clicked.connect(lambda: self.odswiez())
        self.Przycisk.clicked.connect(lambda: self.wyniki())
        self.tabWidget.tabBarDoubleClicked.connect(lambda index: self.set_text(index))
        self.actionOtw_rz.triggered.connect(lambda: self.load_data()) 
        self.actionZapisz.triggered.connect(lambda: self.zapisz())

        self.tabWidget.setTabText(0, "test")

    def closeTab(self, index):
        reply = QMessageBox.question(self, 'Zamykanie okna'," Na pewno?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            self.tabWidget.removeTab(index)
        else:
            return         
    
    def openWindow(self):
        self.window = QtWidgets.QDialog()
        self.ui = Ui_Ustawienia()
        self.ui.setupUi(self.window)
        self.window.show()

        self.ui.pushButton.clicked.connect(lambda: self.ustawienia_ok())

    def ustawienia_ok(self):
        V_o = float(self.ui.V_kolby.text())
        self.window.close()    

    def set_text(self, index):
        text = str(QInputDialog.getText(self, "Zmiana nazwy okienka", "Podaj nową nazwę:"))
        text = text[2:]
        text = text[::-1]
        text= text[8::]
        text= text[::-1]
        if text:
            self.tabWidget.setTabText(index,text)

    def dodaj_temperautre(self):
        self.window = QtWidgets.QDialog()
        self.ui = Ui_okienko()
        self.ui.setupUi(self.window)
        self.tabWidget.addTab(self.window, "Nowa temperatura") 

    def zapisz(self):
               
        przyciski = []
        przyciski.append(self.p_atm.text())
        przyciski.append(self.p_atm_2.text())
        przyciski.append(self.p_atm_3.text())

        my_wb = openpyxl.Workbook()
        #my_wb.title ="Ustawienia" Nie działa
        ws1 = my_wb.create_sheet('500oC')
        ws2 = my_wb.create_sheet('450oC')
        ws3 = my_wb.create_sheet('400oC')

        my_sheet = my_wb.active

        i = 0
        while i <=2:
            my_sheet.cell(row = i+1, column = 8).value = przyciski[i]
            i = i +1

        i= 0
        while i <=5:
            my_sheet.cell(row = 1, column = i+1).value = self.tableWidget.item(0,i).text()
            my_sheet.cell(row = 2, column = i+1).value = self.tableWidget.item(1,i).text()
            my_sheet.cell(row = 3, column = i+1).value = self.tableWidget.item(2,i).text()
            i = i+1

        my_sheet= my_wb["500oC"]
        i= 0
        while i <=5:
            my_sheet.cell(row = 1, column = i+1).value = self.tableWidget_2.item(0,i).text()
            my_sheet.cell(row = 2, column = i+1).value = self.tableWidget_2.item(1,i).text()
            my_sheet.cell(row = 3, column = i+1).value = self.tableWidget_2.item(2,i).text()
            my_sheet.cell(row = 4, column = i+1).value = self.tableWidget_2.item(3,i).text()
            my_sheet.cell(row = 5, column = i+1).value = self.tableWidget_2.item(4,i).text()
            i = i+1
        
        my_sheet= my_wb["450oC"]

        i= 0
        while i <=5:
            my_sheet.cell(row = 1, column = i+1).value = self.tableWidget_3.item(0,i).text()
            my_sheet.cell(row = 2, column = i+1).value = self.tableWidget_3.item(1,i).text()
            my_sheet.cell(row = 3, column = i+1).value = self.tableWidget_3.item(2,i).text()
            my_sheet.cell(row = 4, column = i+1).value = self.tableWidget_3.item(3,i).text()
            my_sheet.cell(row = 5, column = i+1).value = self.tableWidget_3.item(4,i).text()
            i = i+1
        
        my_sheet= my_wb["400oC"]

        i= 0
        while i <=5:
            my_sheet.cell(row = 1, column = i+1).value = self.tableWidget_4.item(0,i).text()
            my_sheet.cell(row = 2, column = i+1).value = self.tableWidget_4.item(1,i).text()
            my_sheet.cell(row = 3, column = i+1).value = self.tableWidget_4.item(2,i).text()
            my_sheet.cell(row = 4, column = i+1).value = self.tableWidget_4.item(3,i).text()
            my_sheet.cell(row = 5, column = i+1).value = self.tableWidget_4.item(4,i).text()
            i = i+1

        filtr = "Roszerzenie (*xlsx);; Pliki Excela (*xlsx)"
        name = QtWidgets.QFileDialog.getSaveFileName(self, 'Otwórz plik', filter= filtr , initialFilter= 'Excel file (*xlsx')
        my_wb.save(name[0]+".xlsx")
    
    
    def load_data(self):
        filtr = "Roszerzenie (*xlsx);; Pliki Excela (*xlsx)"
        name = QtWidgets.QFileDialog.getOpenFileName(self, 'Otwórz plik', filter= filtr , initialFilter= 'Excel file (*xlsx')
        workbook = openpyxl.load_workbook(name[0])
        zakladki = workbook.get_sheet_names()
        sheet = workbook.get_sheet_by_name(zakladki[0])
        list_values = list(sheet.values)
        
        self.p_atm.setText(str(sheet['H1'].value ))
        self.p_atm_2.setText(str(sheet['H2'].value ))
        self.p_atm_3.setText(str(sheet['H3'].value ))

        row_index = 0
        for valu_tuple in list_values:
            col_index = 0
            for value in valu_tuple:
                self.tableWidget.setItem(row_index, col_index, QTableWidgetItem(str(value)))
                col_index = col_index +1
            row_index = row_index +1

        sheet = workbook.get_sheet_by_name(zakladki[1])
        list_values = list(sheet.values)

        row_index = 0
        for valu_tuple in list_values:
            col_index = 0
            for value in valu_tuple:
                self.tableWidget_2.setItem(row_index, col_index, QTableWidgetItem(str(value)))
                col_index = col_index +1
            row_index = row_index +1
        
        sheet = workbook.get_sheet_by_name(zakladki[2])
        list_values = list(sheet.values)

        row_index = 0
        for valu_tuple in list_values:
            col_index = 0
            for value in valu_tuple:
                self.tableWidget_3.setItem(row_index, col_index, QTableWidgetItem(str(value)))
                col_index = col_index +1
            row_index = row_index +1

        sheet = workbook.get_sheet_by_name(zakladki[3])
        list_values = list(sheet.values)

        row_index = 0
        for valu_tuple in list_values:
            col_index = 0
            for value in valu_tuple:
                self.tableWidget_4.setItem(row_index, col_index, QTableWidgetItem(str(value)))
                col_index = col_index +1
            row_index = row_index +1
        
        
    def odswiez(self):
        tabele = []
        tabele.append(self.tableWidget_2)
        tabele.append(self.tableWidget_3)
        tabele.append(self.tableWidget_4)
        i = 0
        z= 0
        while i <= 2:
            while z <=5:
                tabele[i].horizontalHeaderItem(z).setText(self.tableWidget.item(0,z).text())
                z= z +1
            z = 0
            i = i+1

    def wyniki(self):
        self.window = QtWidgets.QDialog()
        self.uiustawienia = Ui_Ustawienia()
        self.uiustawienia.setupUi(self.window)
        V_o = 1.15
        V_o = float(self.uiustawienia.V_kolby.text())
        T_K = 273.15
        V_NH3_wszystkich = []
        T_R_wszystkich = []
        k_wszystkich = []
        k_stosunekxD= []
        Energia_aktywacji = []
        log_k=[]

        z=1
        while z <= 6:
            x= z-1
            t_w = float(self.tableWidget_2.item(3,x).text())
            Cp = float(self.tableWidget_2.item(4,x).text())
            T_w = float(self.tableWidget_2.item(2,x).text())
            p_atm = float(self.p_atm.text())
            T_r = float(self.tableWidget_2.item(0,x).text())
            V_NH3 = (V_o / t_w)*(1+Cp/100)/(1-Cp/100)*3600* T_K/(T_K + T_w) * p_atm/ 760
            T_R = 23.656*T_r + 61.798
            T_R = round(T_R, 1)
            T_R = str(T_R)
            T_R_wszystkich.append(T_R)
            V_NH3 = round(V_NH3, 2)
            V_NH3 = str(V_NH3)
            V_NH3_wszystkich.append(V_NH3)
            z = z+1

        z = 0
        while z <=5:
            x = z-1
            Cp = float(self.tableWidget_2.item(4,z).text())
            Cp= Cp/100
            m = float(self.tableWidget.item(1,z).text())
            k_wyl = float(V_NH3_wszystkich[z])*(Cp/((1+Cp)*m)*(17.03/22.08))
            k_wyl = k_wyl/(float(self.tableWidget.item(2,z).text())/100)
            k_wyl = round(k_wyl, 4)
            k_wszystkich.append(k_wyl)
            stosunek = round(k_wyl/k_wszystkich[0], 4)
            k_stosunekxD.append(stosunek)
            log_szybkosc = math.log(k_wyl)
            log_k.append(log_szybkosc)
            z= z+1                

        z = 1    
        while z <= 6:
            x= z-1
            t_w = float(self.tableWidget_3.item(3,x).text())
            Cp = float(self.tableWidget_3.item(4,x).text())
            T_w = float(self.tableWidget_3.item(2,x).text())
            p_atm = float(self.p_atm_2.text())
            T_r = float(self.tableWidget_3.item(0,x).text())
            V_NH3 = (V_o / t_w)*(1+Cp/100)/(1-Cp/100)*3600* T_K/(T_K + T_w) * p_atm/ 760
            T_R = 23.656*T_r + 61.798
            T_R = round(T_R, 1)
            T_R = str(T_R)
            T_R_wszystkich.append(T_R)
            V_NH3 = round(V_NH3, 2)
            V_NH3 = str(V_NH3)
            V_NH3_wszystkich.append(V_NH3)
            z = z+1

        z = 0
        while z <=5:
            x = z-1
            Cp = float(self.tableWidget_3.item(4,z).text())
            Cp= Cp/100
            m = float(self.tableWidget.item(1,z).text())
            k_wyl = float(V_NH3_wszystkich[z+6])*(Cp/((1+Cp)*m)*(17.03/22.08))
            k_wyl = k_wyl/(float(self.tableWidget.item(2,z).text())/100)
            k_wyl = round(k_wyl, 4)
            k_wszystkich.append(k_wyl)
            stosunek = round(k_wyl/k_wszystkich[6], 4)
            k_stosunekxD.append(stosunek)
            log_szybkosc = math.log(k_wyl)
            log_k.append(log_szybkosc)
            z= z+1   

        z = 1    
        while z <= 6:
            x= z-1
            t_w = float(self.tableWidget_4.item(3,x).text())
            Cp = float(self.tableWidget_4.item(4,x).text())
            T_w = float(self.tableWidget_4.item(2,x).text())
            p_atm = float(self.p_atm_3.text())
            T_r = float(self.tableWidget_4.item(0,x).text())
            V_NH3 = (V_o / t_w)*(1+Cp/100)/(1-Cp/100)*3600* T_K/(T_K + T_w) * p_atm/ 760
            T_R = 23.656*T_r + 61.798
            T_R = round(T_R, 1)
            T_R = str(T_R)
            T_R_wszystkich.append(T_R)
            V_NH3 = round(V_NH3, 2)
            V_NH3 = str(V_NH3)
            V_NH3_wszystkich.append(V_NH3)
            z = z+1
    
        z = 0
        while z <=5:
            Cp = float(self.tableWidget_4.item(4,z).text())
            Cp= Cp/100
            m = float(self.tableWidget.item(1,z).text())
            k_wyl = float(V_NH3_wszystkich[z+12])*(Cp/((1+Cp)*m)*(17.03/22.08))
            k_wyl = k_wyl/(float(self.tableWidget.item(2,z).text())/100)
            k_wyl = round(k_wyl, 4)
            k_wszystkich.append(k_wyl)
            stosunek = round(k_wyl/k_wszystkich[12], 4)
            k_stosunekxD.append(stosunek)
            log_szybkosc = math.log(k_wyl)
            log_k.append(log_szybkosc)
            z= z+1   


        self.window = QtWidgets.QDialog()
        self.ui = Ui_Wyniki()
        self.ui.setupUi(self.window)
        self.window.show()

        self.ui.tabWidget.setTabText(0, "Podsumowanie")
        self.ui.tabWidget.setTabText(1, self.tabWidget.tabText(1))
        self.ui.tabWidget.setTabText(2, self.tabWidget.tabText(2))
        self.ui.tabWidget.setTabText(3, self.tabWidget.tabText(3))
        
        i=0
        os_X= []
        while i <= 17:
            T_R = (float(T_R_wszystkich[i]) +273.15)
            odwr_T_R= 1/T_R
            os_X.append(odwr_T_R)
            i = i+1
        i=0
        while i <=5:
            x = np.array([os_X[i],os_X[i+6],os_X[i+12]])
            y = np.array([log_k[i],log_k[i+6],log_k[i+12]])
            a,b = np.polyfit(x, y, 1)
            plt.scatter(x, y)
            plt.plot(x, a*x+b, linestyle='--', linewidth=2, label = self.tableWidget.item(0,i).text())

            legend = plt.legend(loc= 3, shadow=True, fontsize='x-large')
            legend.get_frame()

            i = i+1
            E_a = round(a*8.314*(-1)/1000, 2)
            Energia_aktywacji.append(E_a)

        plt.show()

        z=0
        while z <= 5:
            self.ui.Tabela_wynikow.verticalHeaderItem(z).setText((self.tableWidget.item(0,z).text()))
            self.ui.Tabela_wynikow.setItem(0, z, QtWidgets.QTableWidgetItem(str(Energia_aktywacji[z])))
            z = z +1

        z=0
        while z <= 5:
            self.ui.Tabela_wynikow_2.verticalHeaderItem(z).setText((self.tableWidget.item(0,z).text()))
            self.ui.Tabela_wynikow_2.setItem(z, 0, QtWidgets.QTableWidgetItem(V_NH3_wszystkich[z]))
            self.ui.Tabela_wynikow_2.setItem(z, 1, QtWidgets.QTableWidgetItem(T_R_wszystkich[z]))
            self.ui.Tabela_wynikow_2.setItem(z, 2, QtWidgets.QTableWidgetItem(self.tableWidget_2.item(1,z).text()))
            self.ui.Tabela_wynikow_2.setItem(z, 3, QtWidgets.QTableWidgetItem(self.tableWidget.item(1,z).text()))
            self.ui.Tabela_wynikow_2.setItem(z, 4, QtWidgets.QTableWidgetItem(str(k_wszystkich[z])))
            self.ui.Tabela_wynikow_2.setItem(z, 5, QtWidgets.QTableWidgetItem(str(k_stosunekxD[z])))
            z = z+1

        z=0
        while z <= 5:
            self.ui.Tabela_wynikow_3.verticalHeaderItem(z).setText((self.tableWidget.item(0,z).text()))
            self.ui.Tabela_wynikow_3.setItem(z, 0, QtWidgets.QTableWidgetItem(V_NH3_wszystkich[z+6]))
            self.ui.Tabela_wynikow_3.setItem(z, 1, QtWidgets.QTableWidgetItem(T_R_wszystkich[z+6]))
            self.ui.Tabela_wynikow_3.setItem(z, 2, QtWidgets.QTableWidgetItem(self.tableWidget_3.item(1,z).text()))
            self.ui.Tabela_wynikow_3.setItem(z, 3, QtWidgets.QTableWidgetItem(self.tableWidget.item(1,z).text()))
            self.ui.Tabela_wynikow_3.setItem(z, 4, QtWidgets.QTableWidgetItem(str(k_wszystkich[z+6])))
            self.ui.Tabela_wynikow_3.setItem(z, 5, QtWidgets.QTableWidgetItem(str(k_stosunekxD[z +6])))
            z = z+1

        z=0
        while z <= 5:
            self.ui.Tabela_wynikow_4.verticalHeaderItem(z).setText((self.tableWidget.item(0,z).text()))
            self.ui.Tabela_wynikow_4.setItem(z, 0, QtWidgets.QTableWidgetItem(V_NH3_wszystkich[z+12]))
            self.ui.Tabela_wynikow_4.setItem(z, 1, QtWidgets.QTableWidgetItem(T_R_wszystkich[z+12]))
            self.ui.Tabela_wynikow_4.setItem(z, 2, QtWidgets.QTableWidgetItem(self.tableWidget_4.item(1,z).text()))
            self.ui.Tabela_wynikow_4.setItem(z, 3, QtWidgets.QTableWidgetItem(self.tableWidget.item(1,z).text()))
            self.ui.Tabela_wynikow_4.setItem(z, 4, QtWidgets.QTableWidgetItem(str(k_wszystkich[z+12])))
            self.ui.Tabela_wynikow_4.setItem(z, 5, QtWidgets.QTableWidgetItem(str(k_stosunekxD[z+12])))
            z = z+1
        
if __name__ == "__main__":
    app = QApplication(sys.argv)
    ui = MainUI()
    ui.show()
    sys.exit(app.exec())
